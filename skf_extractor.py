"""
SKF Observer Phoenix — Extrator de Dados para Power BI
=======================================================
Consulta os endpoints da API Observer e da API Machine Viewer (MHV)
para cada unidade e salva arquivos Excel (.xlsx) organizados por tipo.

Estrutura de saída:
    SKF_PowerBI/
    ├── assets/
    ├── points/
    ├── nextgensensor/
    ├── gateways/
    ├── devices/
    ├── alarms/
    ├── mhv_assets/
    └── mhv_workorders/

Execução:
    python skf_extractor.py

Dependências:
    pip install requests pandas openpyxl
"""

import os
import re
import sys
import time
import json
import warnings
import requests
import pandas as pd
from datetime import datetime, timezone
from openpyxl import load_workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter

# Suprime FutureWarning do pandas para não quebrar a barra de progresso
warnings.filterwarnings("ignore", category=FutureWarning)

# ─────────────────────────────────────────────────────────────────────
# CONFIGURAÇÃO — OBSERVER
# ─────────────────────────────────────────────────────────────────────

# (unit_name, observer_url, mhv_short_name | None)
UNITS = {
    "XXXXXXXX":      ("http://services.repcenter.skf.com:00000", "BRAEO00000"),
    "XXXXXXXX":          ("http://services.repcenter.skf.com:00000", "BRAEO00000"),
    "XXXXXXXX":              ("http://services.repcenter.skf.com:00000", "BRAEO00000"),
    "XXXXXXXX": ("http://services.repcenter.skf.com:00000", "BRAEO00000"),
    "XXXXXXXX":       ("http://services.repcenter.skf.com:00000", "BRAEO00000"),
    "XXXXXXXX":          ("http://services.repcenter.skf.com:00000", "BRAEO00000"),          # sem MHV
}

USERNAME       = "user"
PASSWORD       = ""
TIMEOUT        = 30
TOKEN_VALIDITY = 1140
DELAY          = 0.2
OUTPUT_DIR     = "SKF_PowerBI"

OBSERVER_ENDPOINTS = ["assets", "points", "nextgensensor", "gateways", "devices", "alarms"]
MHV_ENDPOINTS      = ["mhv_assets", "mhv_workorders"]

# ─────────────────────────────────────────────────────────────────────
# CONFIGURAÇÃO — MACHINE VIEWER (MHV)
# ─────────────────────────────────────────────────────────────────────

MHV_BASE_URL  = "https://analystapi.repcenter.skf.com"
MHV_API_KEY   = ""   # preenchido interativamente se vazio

# Período fixo de O.S. (conforme especificado)
MHV_DATE_START = "2024-01-01 00:00:00"
MHV_DATE_END   = "2028-12-31 23:59:59"

# ─────────────────────────────────────────────────────────────────────
# CORES DE TERMINAL (ANSI)
# ─────────────────────────────────────────────────────────────────────

class C:
    RESET   = "\033[0m"
    BOLD    = "\033[1m"
    DIM     = "\033[2m"
    # Texto
    WHITE   = "\033[97m"
    CYAN    = "\033[96m"
    GREEN   = "\033[92m"
    YELLOW  = "\033[93m"
    RED     = "\033[91m"
    BLUE    = "\033[94m"
    MAGENTA = "\033[95m"
    GREY    = "\033[90m"
    # Fundo
    BG_BLUE  = "\033[44m"
    BG_GREEN = "\033[42m"
    BG_RED   = "\033[41m"

def _supports_color():
    return hasattr(sys.stdout, "isatty") and sys.stdout.isatty()

USE_COLOR = _supports_color()

def c(color, text):
    return f"{color}{text}{C.RESET}" if USE_COLOR else text

# ─────────────────────────────────────────────────────────────────────
# INTERFACE DE TERMINAL
# ─────────────────────────────────────────────────────────────────────

WIDTH = 72

def header():
    print()
    print(c(C.BOLD + C.WHITE, "═" * WIDTH))
    print(c(C.BOLD + C.WHITE,
            f"  SKF Observer Phoenix — Extrator de Dados Power BI".center(WIDTH)))
    print(c(C.BOLD + C.WHITE, "═" * WIDTH))
    print(c(C.GREY, f"  {datetime.now().strftime('%d/%m/%Y  %H:%M:%S')}".ljust(WIDTH)))
    print()

def section(title: str):
    print()
    print(c(C.BOLD + C.CYAN, f"  ┌─ {title} " + "─" * max(0, WIDTH - len(title) - 6) + "┐"))

def section_end():
    print(c(C.CYAN, f"  └" + "─" * (WIDTH - 2) + "┘"))

def log_info(msg: str, indent: int = 4):
    prefix = " " * indent
    print(f"{prefix}{c(C.GREY, '·')} {msg}")

def log_ok(msg: str, indent: int = 4):
    prefix = " " * indent
    print(f"{prefix}{c(C.GREEN, '✓')} {msg}")

def log_warn(msg: str, indent: int = 4):
    prefix = " " * indent
    print(f"{prefix}{c(C.YELLOW, '⚠')} {msg}")

def log_err(msg: str, indent: int = 4):
    prefix = " " * indent
    print(f"{prefix}{c(C.RED, '✗')} {msg}")

def log_step(step: str, detail: str = "", indent: int = 4):
    prefix = " " * indent
    line   = f"{prefix}{c(C.BLUE, '→')} {c(C.BOLD, step)}"
    if detail:
        line += f"  {c(C.GREY, detail)}"
    print(line)

def progress_bar(current: int, total: int, label: str = "", width: int = 36):
    """Barra de progresso inline que sobrescreve a linha."""
    pct   = current / total if total else 0
    filled = int(width * pct)
    bar   = "█" * filled + "░" * (width - filled)
    pct_s = f"{pct*100:5.1f}%"
    line  = f"\r  {c(C.CYAN, bar)} {c(C.BOLD, pct_s)}  {label[:28]:<28}"
    sys.stdout.write(line)
    sys.stdout.flush()

def progress_done(label: str = ""):
    bar = "█" * 36
    sys.stdout.write(f"\r  {c(C.GREEN, bar)} {c(C.GREEN + C.BOLD, '100.0%')}  {label[:28]:<28}\n")
    sys.stdout.flush()

def divider():
    print(c(C.GREY, "  " + "─" * (WIDTH - 2)))

def summary_row(label: str, value: str, status: str = "ok"):
    color = C.GREEN if status == "ok" else (C.YELLOW if status == "warn" else C.RED)
    print(f"    {label:<30} {c(color, value)}")

# ─────────────────────────────────────────────────────────────────────
# AUTENTICAÇÃO
# ─────────────────────────────────────────────────────────────────────

_token_cache: dict = {}   # {unit_name: {token, ts}}

def _get_token(unit_name: str, base_url: str) -> str:
    resp = requests.post(
        f"{base_url}/token",
        headers={"Content-Type": "application/x-www-form-urlencoded"},
        data={"grant_type": "password", "username": USERNAME, "password": PASSWORD},
        timeout=TIMEOUT,
    )
    resp.raise_for_status()
    token = resp.json()["access_token"]
    _token_cache[unit_name] = {"token": token, "ts": time.time()}
    return token

def _ensure_token(unit_name: str, base_url: str) -> str:
    cached = _token_cache.get(unit_name, {})
    if cached and (time.time() - cached["ts"]) < TOKEN_VALIDITY:
        return cached["token"]
    return _get_token(unit_name, base_url)

def _hdrs(unit_name: str, base_url: str) -> dict:
    return {
        "Authorization": f"Bearer {_ensure_token(unit_name, base_url)}",
        "Accept": "application/json",
    }

# ─────────────────────────────────────────────────────────────────────
# HELPERS DE API
# ─────────────────────────────────────────────────────────────────────

def _get(unit_name: str, base_url: str, endpoint: str,
         params: dict = None) -> list | dict | None:
    """GET genérico com tratamento de erros."""
    url = f"{base_url}{endpoint}"
    try:
        resp = requests.get(
            url,
            headers=_hdrs(unit_name, base_url),
            params=params,
            timeout=TIMEOUT,
        )
        if resp.status_code == 204:
            return []
        resp.raise_for_status()
        return resp.json()
    except requests.exceptions.Timeout:
        raise TimeoutError(f"Timeout em {endpoint}")
    except requests.exceptions.HTTPError as e:
        raise RuntimeError(f"HTTP {e.response.status_code} em {endpoint}")
    except Exception as e:
        raise RuntimeError(f"Erro em {endpoint}: {e}")

def _normalize(data) -> list:
    """Normaliza resposta para lista."""
    if isinstance(data, list):
        return data
    if isinstance(data, dict):
        return data.get("value", data.get("items", data.get("data", [])))
    return []

# ─────────────────────────────────────────────────────────────────────
# PARSE DE LOCATION (nextgensensor)
# ─────────────────────────────────────────────────────────────────────

def _parse_location(loc: str) -> dict:
    """Parse do campo Location do nextgensensor para hierarquia LDC."""
    parts = [p.strip() for p in re.split(r"[/\\|>]", loc or "") if p.strip()]
    n = len(parts)
    unidade = parts[0] if n > 0 else "—"

    def _is_area(t):
        return t.lower() == "moagem"

    if n >= 5:
        area    = parts[1] if _is_area(parts[1]) else "—"
        offset  = 1 if _is_area(parts[1]) else 0
        setor       = parts[1 + offset] if n > 1 + offset else "—"
        equipamento = parts[2 + offset] if n > 2 + offset else "—"
        ativo       = parts[3 + offset] if n > 3 + offset else "—"
    elif n == 4:
        if _is_area(parts[1]):
            area, setor, equipamento, ativo = parts[1], parts[2], parts[3], "—"
        else:
            area, setor, equipamento, ativo = "—", parts[1], parts[2], parts[3]
    elif n == 3:
        area, setor, equipamento, ativo = "—", parts[1], parts[2], "—"
    else:
        area = "—"
        setor = parts[1] if n > 1 else "—"
        equipamento = "—"
        ativo = "—"

    return {"Unidade": unidade, "Area": area, "Setor": setor,
            "Equipamento": equipamento, "Ativo": ativo}

# ─────────────────────────────────────────────────────────────────────
# EXTRAÇÃO POR ENDPOINT
# ─────────────────────────────────────────────────────────────────────

def fetch_assets(unit_name: str, base_url: str) -> pd.DataFrame:
    data = _normalize(_get(unit_name, base_url, "/v2/assets",
                           params={"includeAcknowledged": "true"}))
    rows = []
    for a in data:
        rows.append({
            "ID":          a.get("ID"),
            "Name":        a.get("Name"),
            "Description": a.get("Description"),
            "Path":        a.get("Path"),
            "Status":      str(a.get("Status", "")),
            "Created":     a.get("SystemCreatedDate") or a.get("Created"),
        })
    return pd.DataFrame(rows)


def fetch_points(unit_name: str, base_url: str,
                 asset_ids: list) -> pd.DataFrame:
    rows = []
    for aid in asset_ids:
        try:
            # Usa /v1/machines/{id}/points — endpoint correto para listar points
            data = _normalize(_get(unit_name, base_url,
                                   f"/v1/machines/{aid}/points"))
        except Exception:
            data = []
        for p in data:
            rows.append({
                "AssetID":   str(aid),
                "PointID":   p.get("ID")       or p.get("id"),
                "Name":      p.get("Name")     or p.get("name"),
                "NodeType":  p.get("NodeType") or p.get("nodeType"),
                "EUType":    p.get("EUType")   or p.get("euType"),
                "Unit":      p.get("Unit")     or p.get("unit"),
                "ParentID":  p.get("ParentID") or p.get("parentId"),
            })
        time.sleep(DELAY)
    return pd.DataFrame(rows)


def fetch_nextgensensor(unit_name: str, base_url: str) -> pd.DataFrame:
    data = _normalize(_get(unit_name, base_url, "/v1/nextgensensor"))
    today = datetime.now(timezone.utc)
    rows = []
    for s in data:
        comm = (s.get("Commissioned") if s.get("Commissioned") is not None
                else s.get("commissioned"))

        conn_raw = s.get("ConnectionState") or s.get("connectionState")
        try:    conn_code = int(conn_raw) if conn_raw is not None else None
        except: conn_code = None
        CONN_LABELS = {0:"Desconectado", 1:"Conectado",
                       2:"Sem Medição", 3:"Conectado — Sem Medição"}
        conn_lbl = CONN_LABELS.get(conn_code, f"? ({conn_raw})")

        diag = s.get("DiagnosticCode") or 0
        try:    diag = int(diag)
        except: diag = 0
        diag_flags = []
        if diag & 1:   diag_flags.append("Bateria Baixa")
        if diag & 512: diag_flags.append("Instabilidade de Rede")

        updated_raw = s.get("StatusLastUpdated") or s.get("statusLastUpdated")
        try:
            updated = pd.to_datetime(updated_raw, utc=True)
            dias_off = (today - updated).days
        except Exception:
            updated  = None
            dias_off = None

        cleared_raw = s.get("ClearedDate") or s.get("clearedDate")
        try:
            cleared = pd.to_datetime(cleared_raw, utc=True)
            cleared = None if cleared.year <= 1940 else cleared
        except Exception:
            cleared = None

        bat = s.get("BatteryLevel") or s.get("batteryLevel")
        try:    bat = float(bat)
        except: bat = None

        loc = _parse_location(s.get("Location") or s.get("location") or "")

        rows.append({
            "IDNode":           s.get("IDNode") or s.get("idNode"),
            "SensorIdentifier": s.get("SensorIdentifier") or "—",
            "IDSmartGateway":   s.get("IDSmartGateway") or s.get("idSmartGateway"),
            "Commissioned":     int(bool(comm)),
            "BatteryLevel":     bat,
            "ConnectionState":  conn_lbl,
            "DiagnosticCode":   diag,
            "DiagFlags":        ", ".join(diag_flags) if diag_flags else "—",
            "StatusLastUpdated":updated_raw,
            "DiasOffline":      dias_off,
            "ClearedDate":      cleared.strftime("%Y-%m-%d %H:%M:%S") if cleared else None,
            "FWVersion":        s.get("FWVersion"),
            "HWVersion":        s.get("HWVersion"),
            "Temperature":      s.get("Temperature"),
            "Name":             s.get("Name") or s.get("name"),
            **loc,
        })
    return pd.DataFrame(rows)


def fetch_gateways(unit_name: str, base_url: str) -> pd.DataFrame:
    data = _normalize(_get(unit_name, base_url, "/v1/gateways"))
    today = datetime.now(timezone.utc)
    rows = []
    for g in data:
        updated_raw = g.get("statusLastUpdated") or g.get("StatusLastUpdated")
        try:
            updated  = pd.to_datetime(updated_raw, utc=True)
            dias_off = (today - updated).days
        except Exception:
            updated  = None
            dias_off = None
        connected = g.get("connected") if g.get("connected") is not None else g.get("Connected")
        rows.append({
            "GatewayID":         g.get("id") or g.get("ID"),
            "Name":              g.get("name") or g.get("Name"),
            "HardwareId":        g.get("hardwareId") or g.get("HardwareId"),
            "Connected":         int(bool(connected)),
            "Status":            "Online" if connected else "Offline",
            "StatusLastUpdated": updated_raw,
            "DiasSemUpdate":     dias_off,
        })
    return pd.DataFrame(rows)


def fetch_devices(unit_name: str, base_url: str) -> pd.DataFrame:
    data = _normalize(_get(unit_name, base_url, "/v1/device"))
    today = datetime.now(timezone.utc)
    SYNC_LABELS = {0:"Não Sincronizado", 1:"Sincronizado",
                   2:"Pendente", 100:"Falha"}
    rows = []
    for d in data:
        updated_raw = d.get("lastupdate") or d.get("LastUpdate")
        try:
            updated  = pd.to_datetime(updated_raw, utc=True)
            dias_off = (today - updated).days
        except Exception:
            updated  = None
            dias_off = None
        sync_code = d.get("synchronizationstatus")
        try:    sync_code = int(sync_code)
        except: sync_code = None
        rows.append({
            "DeviceID":    d.get("id") or d.get("ID"),
            "Name":        d.get("name") or d.get("Name"),
            "Active":      int(bool(d.get("active") or d.get("Active"))),
            "LastUpdate":  updated_raw,
            "DiasOffline": dias_off,
            "SyncCode":    sync_code,
            "SyncStatus":  SYNC_LABELS.get(sync_code, f"? ({sync_code})"),
        })
    return pd.DataFrame(rows)


def fetch_alarms(unit_name: str, base_url: str) -> pd.DataFrame:
    data = _normalize(_get(unit_name, base_url, "/v2/alarms"))

    def _scalar(v):
        """Converte dicts/listas aninhados para string JSON — o Excel não aceita objetos."""
        if isinstance(v, (dict, list)):
            return json.dumps(v, ensure_ascii=False)
        return v

    rows = []
    for a in data:
        rows.append({
            "AlarmID":      _scalar(a.get("id")          or a.get("ID")),
            "PointID":      _scalar(a.get("pointId")     or a.get("PointId")  or a.get("PointID")),
            "Type":         _scalar(a.get("type")        or a.get("Type")),
            "Severity":     _scalar(a.get("severity")    or a.get("Severity")),
            "Value":        _scalar(a.get("value")       or a.get("Value")),
            "Limit":        _scalar(a.get("limit")       or a.get("Limit")),
            "Timestamp":    _scalar(a.get("timestamp")   or a.get("Timestamp") or a.get("date")),
            "Acknowledged": int(bool(a.get("acknowledged") or a.get("Acknowledged"))),
            "Message":      _scalar(a.get("message")     or a.get("Message") or a.get("description")),
            "Trigger":      _scalar(a.get("trigger")     or a.get("Trigger")),
        })
    return pd.DataFrame(rows)

# ─────────────────────────────────────────────────────────────────────
# EXPORTAÇÃO XLSX
# ─────────────────────────────────────────────────────────────────────

# Cores LDC para o Excel
XL_HEADER_BG   = "1E3A4C"   # LDC Blue escuro
XL_HEADER_FONT = "FFFFFF"
XL_ALT_ROW     = "EAF1F6"   # azul muito claro
XL_BORDER_CLR  = "C5D5E0"
XL_META_BG     = "F0F3F6"
XL_META_FONT   = "5C6670"


def _xl_border(color: str = XL_BORDER_CLR) -> Border:
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _autofit(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)


def save_xlsx(df: pd.DataFrame, filepath: str,
              unit_name: str, endpoint: str, ts: str):
    """Salva DataFrame como .xlsx com formatação LDC profissional."""
    from openpyxl import Workbook

    # Aplaina qualquer valor dict/lista remanescente (proteção geral)
    def _flat(v):
        if isinstance(v, (dict, list)):
            return json.dumps(v, ensure_ascii=False)
        return v

    df = df.map(_flat)

    wb = Workbook()
    ws = wb.active
    ws.title = endpoint[:31]

    # ── Meta info (linha 1) ───────────────────────────────────────
    ws.merge_cells("A1:H1")
    meta_cell = ws["A1"]
    meta_cell.value = (f"SKF Observer Phoenix  ·  {unit_name.replace('_',' ')}  ·  "
                       f"{endpoint.upper()}  ·  Atualizado: {ts}")
    meta_cell.font      = Font(name="Arial", size=9, color=XL_META_FONT)
    meta_cell.fill      = PatternFill("solid", fgColor=XL_META_BG)
    meta_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 18

    # ── Cabeçalho (linha 2) ───────────────────────────────────────
    header_fill = PatternFill("solid", fgColor=XL_HEADER_BG)
    header_font = Font(name="Arial", size=10, bold=True, color=XL_HEADER_FONT)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=2, column=col_idx, value=col_name)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = header_align
        cell.border    = _xl_border()
    ws.row_dimensions[2].height = 22

    # ── Dados (linha 3 em diante) ─────────────────────────────────
    alt_fill  = PatternFill("solid", fgColor=XL_ALT_ROW)
    data_font = Font(name="Arial", size=9)
    data_border = _xl_border()

    for row_idx, (_, row) in enumerate(df.iterrows(), start=3):
        fill = alt_fill if row_idx % 2 == 0 else None
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font      = data_font
            cell.border    = data_border
            cell.alignment = Alignment(vertical="center")
            if fill:
                cell.fill = fill

    # ── Freeze panes + autofilter ─────────────────────────────────
    ws.freeze_panes = "A3"
    if not df.empty:
        last_col = get_column_letter(len(df.columns))
        ws.auto_filter.ref = f"A2:{last_col}{len(df) + 2}"

    # ── Autofit colunas ───────────────────────────────────────────
    _autofit(ws)
    ws.row_dimensions[2].height = 22   # restaura após autofit

    wb.save(filepath)

# ─────────────────────────────────────────────────────────────────────
# MACHINE VIEWER (MHV) — FUNÇÕES DE FETCH
# ─────────────────────────────────────────────────────────────────────

def _mhv_post(short_name: str, endpoint: str, payload: dict) -> dict | None:
    """POST genérico para a API Machine Viewer com autenticação via x-api-key."""
    url = f"{MHV_BASE_URL}/{short_name}/{endpoint}"
    try:
        resp = requests.post(
            url,
            json=payload,
            headers={
                "x-api-key":    MHV_API_KEY,
                "Content-Type": "application/json",
                "Accept":       "application/json",
            },
            timeout=TIMEOUT,
        )
        if resp.status_code == 204:
            return None
        resp.raise_for_status()
        return resp.json()
    except requests.exceptions.Timeout:
        raise TimeoutError(f"Timeout em MHV /{endpoint}")
    except requests.exceptions.HTTPError as e:
        body = ""
        try:
            body = e.response.json().get("message", e.response.text[:200])
        except Exception:
            body = e.response.text[:200]
        raise RuntimeError(f"HTTP {e.response.status_code} em MHV /{endpoint}: {body}")
    except Exception as e:
        raise RuntimeError(f"Erro em MHV /{endpoint}: {e}")


def _scalar_mhv(v):
    """Converte dicts/listas para string JSON — o Excel não aceita objetos."""
    if isinstance(v, (dict, list)):
        return json.dumps(v, ensure_ascii=False)
    return v


def mhv_fetch_assets(short_name: str) -> pd.DataFrame:
    """
    POST /{short}/assets — lista completa de ativos com paginação via nextCursor.
    """
    expression_base = (
        "{ assetId assetName assetDescription functionalLocation "
        "parentId parentName criticality assetStatus assetSegment conditionIndex }"
    )

    all_rows = []
    cursor   = None
    page     = 0

    while True:
        page += 1
        if cursor:
            expr = f'{{ filter(cursor: {cursor}) {expression_base} }}'
        else:
            expr = f'{{ filter {expression_base} }}'

        payload = {"expression": expr}
        data    = _mhv_post(short_name, "assets", payload)
        if not data:
            break

        items = data.get("data", [])
        for a in items:
            all_rows.append({
                "AssetId":           _scalar_mhv(a.get("assetId")),
                "AssetName":         _scalar_mhv(a.get("assetName")),
                "AssetDescription":  _scalar_mhv(a.get("assetDescription")),
                "FunctionalLocation":_scalar_mhv(a.get("functionalLocation")),
                "ParentId":          _scalar_mhv(a.get("parentId")),
                "ParentName":        _scalar_mhv(a.get("parentName")),
                "Criticality":       _scalar_mhv(a.get("criticality")),
                "AssetStatus":       _scalar_mhv(a.get("assetStatus")),
                "AssetSegment":      _scalar_mhv(a.get("assetSegment")),
                "ConditionIndex":    _scalar_mhv(a.get("conditionIndex")),
            })

        cursor = data.get("nextCursor")
        if not cursor:
            break
        time.sleep(DELAY)

    return pd.DataFrame(all_rows)


def mhv_fetch_workorders(short_name: str) -> pd.DataFrame:
    """
    POST /{short}/workorders — ordens de serviço no período fixo com paginação.
    Período: 2024-01-01 00:00:00 → 2028-12-31 23:59:59
    """
    all_rows = []
    cursor   = None
    page     = 0

    fields = (
        "assetId id orderNumber deadline priority technique "
        "scheduledDate openingDate reWork cmmsRegister cmms "
        "services situation author "
        "intervention{ date interventionType description isDiagnosticCorrect }"
    )

    while True:
        page += 1
        if cursor:
            expr = (
                f'{{ filter('
                f'openingDateStart: "{MHV_DATE_START}", '
                f'openingDateEnd: "{MHV_DATE_END}", '
                f'cursor: {cursor}'
                f') {{ {fields} }} }}'
            )
        else:
            expr = (
                f'{{ filter('
                f'openingDateStart: "{MHV_DATE_START}", '
                f'openingDateEnd: "{MHV_DATE_END}"'
                f') {{ {fields} }} }}'
            )

        payload = {"expression": expr}
        data    = _mhv_post(short_name, "workorders", payload)
        if not data:
            break

        items = data.get("data", [])
        for w in items:
            interv = w.get("intervention") or {}
            all_rows.append({
                "AssetId":          _scalar_mhv(w.get("assetId")),
                "Id":               _scalar_mhv(w.get("id")),
                "OrderNumber":      _scalar_mhv(w.get("orderNumber")),
                "Deadline":         _scalar_mhv(w.get("deadline")),
                "Priority":         _scalar_mhv(w.get("priority")),
                "Technique":        _scalar_mhv(w.get("technique")),
                "ScheduledDate":    _scalar_mhv(w.get("scheduledDate")),
                "OpeningDate":      _scalar_mhv(w.get("openingDate")),
                "ReWork":           _scalar_mhv(w.get("reWork")),
                "CmmsRegister":     _scalar_mhv(w.get("cmmsRegister")),
                "Cmms":             _scalar_mhv(w.get("cmms")),
                "Services":         _scalar_mhv(w.get("services")),
                "Situation":        _scalar_mhv(w.get("situation")),
                "Author":           _scalar_mhv(w.get("author")),
                "Interv_Date":      _scalar_mhv(interv.get("date")),
                "Interv_Type":      _scalar_mhv(interv.get("interventionType")),
                "Interv_Desc":      _scalar_mhv(interv.get("description")),
                "Interv_DiagOK":    _scalar_mhv(interv.get("isDiagnosticCorrect")),
            })

        cursor = data.get("nextCursor")
        if not cursor:
            break
        time.sleep(DELAY)

    return pd.DataFrame(all_rows)


# ─────────────────────────────────────────────────────────────────────
# ORQUESTRAÇÃO POR UNIDADE
# ─────────────────────────────────────────────────────────────────────

FETCH_MAP = {
    "assets":         None,   # tratado separadamente (precisa de asset_ids para points)
    "points":         None,
    "nextgensensor":  fetch_nextgensensor,
    "gateways":       fetch_gateways,
    "devices":        fetch_devices,
    "alarms":         fetch_alarms,
}


def process_unit(unit_name: str, base_url: str, mv_short: str | None,
                 ts: str, results: dict) -> dict:
    """
    Processa uma unidade completa:
      Fase 1 — Observer (assets, points, nextgensensor, gateways, devices, alarms)
      Fase 2 — Machine Viewer (mhv_assets, mhv_workorders) se mv_short disponível
    """
    unit_results = {}

    # ── FASE 1: Observer ─────────────────────────────────────────
    log_step("Observer — Autenticando", base_url)
    try:
        _get_token(unit_name, base_url)
        log_ok("Token obtido")
    except Exception as e:
        log_err(f"Falha de autenticação Observer: {e}")
        for ep in OBSERVER_ENDPOINTS:
            unit_results[ep] = {"rows": 0, "file": None,
                                "status": "err", "msg": "auth failed"}
        # Continua para tentar MHV mesmo se Observer falhar
    else:
        n_obs = len(OBSERVER_ENDPOINTS)
        for ep_idx, endpoint in enumerate(OBSERVER_ENDPOINTS):
            progress_bar(ep_idx, n_obs, endpoint)

            out_dir  = os.path.join(OUTPUT_DIR, endpoint)
            os.makedirs(out_dir, exist_ok=True)
            filename = f"{endpoint}_{unit_name}.xlsx"
            filepath = os.path.join(out_dir, filename)

            try:
                if endpoint == "assets":
                    df = fetch_assets(unit_name, base_url)
                    asset_ids = df["ID"].dropna().tolist()
                    results["_asset_ids"][unit_name] = asset_ids

                elif endpoint == "points":
                    asset_ids = results["_asset_ids"].get(unit_name, [])
                    if not asset_ids:
                        raise RuntimeError("Assets não carregados — points ignorados")
                    df = fetch_points(unit_name, base_url, asset_ids)

                else:
                    df = FETCH_MAP[endpoint](unit_name, base_url)

                if df.empty:
                    unit_results[endpoint] = {"rows": 0, "file": None,
                                              "status": "warn", "msg": "sem dados"}
                else:
                    save_xlsx(df, filepath, unit_name, endpoint, ts)
                    unit_results[endpoint] = {"rows": len(df), "file": filepath,
                                              "status": "ok", "msg": ""}

            except Exception as e:
                unit_results[endpoint] = {"rows": 0, "file": None,
                                          "status": "err", "msg": str(e)}
            time.sleep(DELAY)

        progress_done(f"Observer {unit_name}")

    # ── FASE 2: Machine Viewer ────────────────────────────────────
    if not mv_short:
        log_info("MHV: sem short name para esta unidade — pulando.")
        for ep in MHV_ENDPOINTS:
            unit_results[ep] = {"rows": 0, "file": None,
                                "status": "warn", "msg": "sem MHV nesta unidade"}
        return unit_results

    if not MHV_API_KEY:
        log_warn("MHV: API Key não informada — pulando.")
        for ep in MHV_ENDPOINTS:
            unit_results[ep] = {"rows": 0, "file": None,
                                "status": "warn", "msg": "API Key ausente"}
        return unit_results

    log_step("MHV — Coletando dados", f"{MHV_BASE_URL}/{mv_short}")

    MHV_FETCH = {
        "mhv_assets":     lambda: mhv_fetch_assets(mv_short),
        "mhv_workorders": lambda: mhv_fetch_workorders(mv_short),
    }

    n_mhv = len(MHV_ENDPOINTS)
    for ep_idx, endpoint in enumerate(MHV_ENDPOINTS):
        progress_bar(ep_idx, n_mhv, endpoint)

        out_dir  = os.path.join(OUTPUT_DIR, endpoint)
        os.makedirs(out_dir, exist_ok=True)
        filename = f"{endpoint}_{unit_name}.xlsx"
        filepath = os.path.join(out_dir, filename)

        try:
            df = MHV_FETCH[endpoint]()
            if df.empty:
                unit_results[endpoint] = {"rows": 0, "file": None,
                                          "status": "warn", "msg": "sem dados"}
            else:
                save_xlsx(df, filepath, unit_name, endpoint, ts)
                unit_results[endpoint] = {"rows": len(df), "file": filepath,
                                          "status": "ok", "msg": ""}
        except Exception as e:
            unit_results[endpoint] = {"rows": 0, "file": None,
                                      "status": "err", "msg": str(e)}
        time.sleep(DELAY)

    progress_done(f"MHV {unit_name}")
    return unit_results

# ─────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────

def main():
    global USERNAME, PASSWORD

    header()

    # ── Seleção de unidades ───────────────────────────────────────
    section("Seleção de Unidades")
    units_list = list(UNITS.items())
    print()
    for i, (name, (url, short)) in enumerate(units_list, 1):
        label = name.replace("_", " ")
        mv_tag = c(C.MAGENTA, f"MHV:{short}") if short else c(C.GREY, "sem MHV")
        print(f"    {c(C.CYAN, str(i))}.  {c(C.BOLD, label):<28}  {c(C.GREY, url)}  {mv_tag}")
    print(f"    {c(C.CYAN, '0')}.  {c(C.BOLD, 'TODAS AS UNIDADES')}")
    print()

    raw = input(c(C.WHITE, "  Unidade(s) [0-5, separadas por vírgula]: ")).strip()
    if raw == "0" or raw == "":
        selected = units_list
    else:
        indices  = [int(x.strip()) - 1 for x in raw.split(",")
                    if x.strip().isdigit() and 1 <= int(x.strip()) <= len(units_list)]
        selected = [units_list[i] for i in indices]

    if not selected:
        log_err("Nenhuma unidade válida selecionada. Encerrando.")
        return

    section_end()

    # ── Credenciais ───────────────────────────────────────────────
    section("Credenciais")
    print()
    username_input = input(c(C.WHITE, f"  Usuário [{USERNAME}]: ")).strip()
    if username_input:
        USERNAME = username_input

    log_info(f"Usuário: {c(C.BOLD, USERNAME)}")

    if not PASSWORD:
        print(f"  {c(C.WHITE, 'Senha')} {c(C.GREY, '(visível ao digitar):')}")
        PASSWORD = input(f"  → ").strip()
    else:
        log_info("Senha: carregada do script")

    # MHV API Key (só pede se houver unidades com MHV)
    has_mhv = any(short for _, (_, short) in selected)
    if has_mhv:
        print()
        print(f"  {c(C.MAGENTA, 'API Key — Machine Viewer')} {c(C.GREY, '(visível ao digitar):')}")
        global MHV_API_KEY
        if not MHV_API_KEY:
            MHV_API_KEY = input(f"  → ").strip()
        else:
            log_info(f"MHV API Key: carregada do script")
    section_end()

    # ── Diretório de saída ────────────────────────────────────────
    section("Configuração")
    print()
    ts = datetime.now().strftime("%Y%m%d_%H%M")

    # Cria toda a estrutura de pastas antes de qualquer chamada à API
    try:
        for ep in OBSERVER_ENDPOINTS + MHV_ENDPOINTS:
            os.makedirs(os.path.join(OUTPUT_DIR, ep), exist_ok=True)
        log_ok(f"Diretório de saída: {c(C.BOLD, os.path.abspath(OUTPUT_DIR))}")
    except Exception as e:
        log_err(f"Não foi possível criar o diretório '{OUTPUT_DIR}': {e}")
        return
    log_info(f"Timestamp dos arquivos: {c(C.BOLD, ts)}")
    log_info(f"Unidades selecionadas: {c(C.BOLD, str(len(selected)))}")
    log_info(f"Endpoints Observer: {c(C.BOLD, ', '.join(OBSERVER_ENDPOINTS))}")
    log_info(f"Endpoints MHV:      {c(C.BOLD, ', '.join(MHV_ENDPOINTS))}")
    section_end()

    # ── Processamento escalonado ──────────────────────────────────
    all_results = {"_asset_ids": {}}
    grand_start = time.time()
    total_files = 0
    total_rows  = 0
    errors      = []

    for u_idx, (unit_name, (base_url, mv_short)) in enumerate(selected):
        label = unit_name.replace("_", " ")
        section(f"[{u_idx+1}/{len(selected)}] {label}")
        print()
        t0 = time.time()

        unit_res = process_unit(unit_name, base_url, mv_short, ts, all_results)
        all_results[unit_name] = unit_res

        # Resumo da unidade
        print()
        divider()
        for ep, info in unit_res.items():
            status = info["status"]
            if status == "ok":
                summary_row(ep, f"{info['rows']:>6} linhas  →  {os.path.basename(info['file'])}",
                            "ok")
                total_files += 1
                total_rows  += info["rows"]
            elif status == "warn":
                summary_row(ep, f"sem dados  ({info['msg']})", "warn")
            else:
                summary_row(ep, f"ERRO: {info['msg']}", "err")
                errors.append(f"{unit_name}/{ep}: {info['msg']}")

        elapsed = time.time() - t0
        print()
        log_info(f"Tempo da unidade: {elapsed:.1f}s")
        section_end()

        # Pausa entre unidades (exceto na última)
        if u_idx < len(selected) - 1:
            print()
            log_info(f"Aguardando 3s antes da próxima unidade…")
            time.sleep(3)

    # ── Resumo final ──────────────────────────────────────────────
    elapsed_total = time.time() - grand_start
    print()
    print(c(C.BOLD + C.WHITE, "═" * WIDTH))
    print(c(C.BOLD + C.WHITE, "  RESUMO FINAL".center(WIDTH)))
    print(c(C.BOLD + C.WHITE, "═" * WIDTH))
    print()
    summary_row("Unidades processadas", str(len(selected)), "ok")
    summary_row("Arquivos gerados",     str(total_files),   "ok" if total_files else "warn")
    summary_row("Total de linhas",      f"{total_rows:,}",  "ok")
    summary_row("Erros",               str(len(errors)),    "ok" if not errors else "err")
    summary_row("Tempo total",          f"{elapsed_total:.1f}s", "ok")
    summary_row("Saída",               os.path.abspath(OUTPUT_DIR), "ok")

    if errors:
        print()
        print(c(C.YELLOW, "  Erros encontrados:"))
        for e in errors:
            log_err(e)

    print()
    print(c(C.BOLD + C.GREEN, "  Extração concluída.".center(WIDTH)))
    print(c(C.BOLD + C.WHITE, "═" * WIDTH))
    print()


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print()
        print()
        log_warn("Interrompido pelo usuário (Ctrl+C).")
        print()
        sys.exit(0)
